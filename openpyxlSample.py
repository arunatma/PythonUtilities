# References:
# https://automatetheboringstuff.com/chapter12/
# https://openpyxl.readthedocs.io

import openpyxl as xl
# from config import *
import re
import pprint
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font
from openpyxl import chart as xlct
import random

def get_cell_range(start_cell, num_row = 1, num_col = 1):
    start_cell = start_cell.upper()
    m = re.match(r"([a-zA-Z]+)([0-9]+)", start_cell)
    letter_value = m.group(1)
    number_value = m.group(2)
    start_addr = "$" + letter_value + "$" + number_value
    letter_value = get_column_letter(
        column_index_from_string(letter_value) + num_col - 1)
    number_value = str(eval(number_value) + num_row - 1)
    end_addr = "$" + letter_value + "$" + number_value
    ref_addr = start_addr + ":" + end_addr
    return ref_addr


# Reading items from an excel sheet 
workBook = xl.load_workbook('teams.xlsx')

# Get all the sheet names
sheets = workBook.sheetnames
sheet1 = workBook[sheets[0]]

# Properties of sheet1
title1 = sheet1.title


# To get column letter from number: get_column_letter 
# To get column number from letter: column_index_from_string
# Both of the above functions are in openpyxl.utils, used in get_cell_range
last_cell = get_column_letter(sheet1.max_column) + str(sheet1.max_row)
cell_range = 'A1' + ":" + last_cell

# Get all values
for row in sheet1[cell_range]:
    for cell in row:
        print(cell.coordinate, cell.value)
    print ("------")

# Load values in the desired data structure    
pointsTable = {}
data_range = 'A2' + ":" + last_cell
for row in sheet1[data_range]:
    group   = row[0].value
    team    = row[1].value
    played  = row[2].value
    won     = row[3].value
    draw    = row[4].value
    lost    = row[5].value
    if not group in pointsTable.keys():
        pointsTable[group] = {}
    pointsTable[group][team] = {
        "played" : played, 
        "won" : won,
        "draw" : draw,
        "lost" : lost
#       "points" : won * 3 + draw * 1 + lost * 0
    }

# Calculating Points
for group in pointsTable.keys():
    for team in pointsTable[group].keys():
        won  = pointsTable[group][team]["won"] 
        draw = pointsTable[group][team]["draw"]
        pointsTable[group][team]["points"] = won * 3 + draw * 1
        
        
# Writing a few items on to the excel sheet and saving it
pointsFile = open('fifaPoints.py', 'w')
pointsFile.write('pointsTable = ' + pprint.pformat(pointsTable))
pointsFile.close()

# Calculating Points via Excel
sheet1['G1'] = 'Points'
for i in range(2, sheet1.max_row + 1):
    j = str(i)
    sheet1['G'+j] = '=D'+j + '*3+' + 'E'+j + '*1'

# Change Font, Increase Size and Italicize and Bolden the headings
newFont = Font(name="Courier New", size=16, bold=True, italic=True)
for row in sheet1['A1:G1']:
    for cell in row:
        cell.font = newFont

# Changing the dimensions
# Let us make the first column a little wider and the first row a little taller

# height: 15 points or 20 pixels   (Points means 1/72 of inch on printable paper)
# width: 8.43 characters or 64 pixels  (standard width is 8.43 characters)
#https://support.microsoft.com/en-in/help/214123/description-of-how-column-widths-are-determined-in-excel
# height eqn: pixels = 4/3 * points 
# width eqn : pixels = 5 + 7 * points

# Takes points as input. To get a square cell of A1 needs a different width 
# and height: To get 96 pixels height and width 
# height points = 3/4 * 100 = 72
# width points = (96 - 5)/7 =  13
sheet1.row_dimensions[1].height = 72
sheet1.column_dimensions['A'].width = 13

# Saving it in a different file     
workBook.save("teamsPoints1.xlsx")

# Adding and Removing Sheets 
workBook.create_sheet(index=0, title='Sheet0')
workBook.sheetnames             #['Sheet0', 'Sheet1']
workBook.create_sheet(index=1, title='Sheet0.5')
workBook.sheetnames             #['Sheet0', 'Sheet0.5', 'Sheet1']

workBook.remove(workBook['Sheet0.5'])
workBook.sheetnames             #['Sheet0', 'Sheet1']
workBook.remove(workBook['Sheet0'])
workBook.sheetnames             #['Sheet1']

# Insert a new row or new colum
# Caution: Insertion here will not do formula update
# Insertion in excel will automatically do the formula update 
sheet1.insert_rows(idx=0, amount=1)
sheet1.insert_cols(idx=0, amount=1)

# Merge and Unmerge
sheet1.merge_cells('B1:H1')
sheet1['B1'] = "Points Table"

sheet1.merge_cells('A2:A10')
sheet1['A2'] = "Groups C and D"

# Unmerge 
sheet1.unmerge_cells('A2:A10')

# Freeze Panes
sheet1.freeze_panes     = 'A2'      # Freeze Row 1 
sheet1.freeze_panes     = 'B1'      # Freeze Column A
sheet1.freeze_panes     = 'D5'      # Freeze Cols A to C and Rows 1 to 4
sheet1.freeze_panes     = 'A1'      # Unfreeze all
sheet1.freeze_panes     = None      # Unfreeze all

# Charting
for i in range(1, 11):              # Create Data in Col L
    sheet1['L' + str(i)] = random.randint(1,11)

theRef = xlct.Reference(sheet1, min_col=12, min_row=1, max_col=12, max_row=10)
theSeries = xlct.Series(theRef, title='Random Values')
theChart = xlct.BarChart()

theChart.title = 'Bar Chart'
theChart.append(theSeries)
sheet1.add_chart(theChart, 'M14')

theChart = xlct.LineChart()
theChart.title = 'Line Chart'
theChart.append(theSeries)
sheet1.add_chart(theChart, 'M30')

workBook.save('sampleChart.xlsx')    
